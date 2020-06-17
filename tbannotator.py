#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jun 12 13:24:26 2020

@authors: Christophe Guyeux, Guislaine Refrégier, Christophe Sola
"""

import argparse
import Bio.Entrez
import Bio.pairwise2
import Bio.SeqIO
import collections
import distutils.spawn
import logging
import logging.config
import networkx as nx
import openpyxl
import os
import os.path
import pathlib
import pickle
import shutil
import stat
import subprocess as sp
import sys
import tarfile
import time
import urllib.request
import xlrd
import xlsxwriter
import xmltodict
import zipfile

  
class TBannotator:
    """
    Main class that provides various informations related to a Sequence
    Read Archive (SRA) accession number of a tuberculosis genome
    """

    def __init__(self):
        # Dictionary of results
        self._results = {}
        # To measure the execution time
        self._t0 = time.time()
        # Parse arguments
        self._parse_args()
        self._linux_or_mac = sys.platform.startswith('linux') or sys.platform.startswith('darwin')
        if not self._linux_or_mac:
            self._logger.warning(f'{sys.platform} OS not recommended (sharp reduction in speed')
        self._logger.info('Initialization')
        self._dir_data = pathlib.Path('data')
        self._read_h37Rv()
        self._read_SIT()
        self._run()


    def _run(self):
        if 'bioproject' in self._results:
            self.__bioproject_to_sra()
        for sra in self.sras:
            try:
                self._check_for_tools()
                self._logger.info(f'SRA = {sra}')
                self._sra = sra
                self._results[self._sra] = {}
                self._prepare_directory()
                self._collect_sra()
                self._prepare_sra()
                self._seq_info()
                #self._ncbi_info()
                self._make_blast_db()
                # Blasts of in silico/vitro spoligotypes
                self._blast_spoligo()
                if self._with_spol_vitro:
                    self._blast_spoligo_vitro()
                # SIT value
                self._SIT()
                # Getting lineages
                if self._with_Coll:
                    self._lineage_Coll()
                if self._with_Coll_full:
                    self._lineage_Coll_full()
                if self._with_L6_animal:
                    self._lineage_L6_animal()
                if self._with_PGG:
                    self._lineage_PGG()
                if self._with_Palittapongarnpim:
                    self._lineage_Palittapongarnpim()
                if self._with_Palittapongarnpim_full:
                    self._lineage_Palittapongarnpim_full()
                if self._with_Shitikov:
                    self._lineage_Shitikov()
                if self._with_Stucki:
                    self._lineage_Stucki()
                if self._with_IS:
                    self._find_IS()
                # Print a summary of results
                self._local_summary()
                self._summary()
            except NotImplementedError:
                pass


    def __str2bool(self, v):
        if isinstance(v, bool):
           return v
        if v.lower() in ('yes', 'true', 't', 'y', '1'):
            return True
        elif v.lower() in ('no', 'false', 'f', 'n', '0'):
            return False
        else:
            raise argparse.ArgumentTypeError('Boolean value expected.')        
        
        
    def _parse_args(self):
        """
        Parses the arguments provided to the TB-profiler
        """
        parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
        parser.add_argument("-sra",
                            help="accession number to deal with",
                            default='',
                            type=str)
        parser.add_argument("-list",
                            "--sra_list",
                            help="SRA list in text file, one per line",
                            default='',
                            type=str)
        parser.add_argument("-bioproject",
                            help="NCBI BioProject ID where to find SRAs",
                            default='',
                            type=str)
        parser.add_argument('-d',
                            '--debug',
                            help="debugging mode of logs",
                            action="store_const",
                            dest="loglevel",
                            const=logging.DEBUG,
                            default=logging.WARNING
                            )
        parser.add_argument('-v',
                            '--verbose',
                            help="verbose logs",
                            action="store_const",
                            dest="loglevel",
                            const=logging.INFO
                            )
        parser.add_argument("-out",
                            "--output_directory",
                            default='sequences',
                            help="directory of outputs",
                            type=str)
        parser.add_argument("-out_dict",
                            "--output_dict",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="store results in a dictionary (pickle)")
        parser.add_argument("--num_threads",
                            default='1',
                            help="number of threads",
                            type=str)
        parser.add_argument("--spol_evalue",
                            default='1e-6',
                            help="evalue when blasting spoligotype spacers",
                            type=str)
        parser.add_argument("--spol_nbmatches",
                            default=5,
                            help="number of blast matches to consider the presence of a spacer",
                            type=int)
        parser.add_argument("-spol_vitro",
                            "--with_spol_vitro",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="simulate spoligotypes in vitro via blast")
        parser.add_argument("--lineage_evalue",
                            default='1e-5',
                            help="evalue when blasting lineage SNPs",
                            type=str)
        parser.add_argument("--lineage_blast_ws",
                            default=20,
                            help="in blast, number of nucleotides (window size) before and after the SNP",
                            type=str)
        parser.add_argument("--lineage_validate_ws",
                            default=5,
                            help="window size of exact matching in blast results of SNP lineage",
                            type=str)
        parser.add_argument("-Coll",
                            "--with_Coll",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=True,
                            help="Compute the Coll et al. lineage")
        parser.add_argument("-Coll_full",
                            "--with_Coll_full",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the Coll et al. lineage (full SNPs)")
        parser.add_argument("-L6_animal",
                            "--with_L6_animal",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the L6+animal lineage")
        parser.add_argument("-PGG",
                            "--with_PGG",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the PGG lineage")
        parser.add_argument("-Pali",
                            "--with_Palittapongarnpim",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the Palittapongarnpim et al. lineage for L1 (one SNP per sublineage)")
        parser.add_argument("-Pali_full",
                            "--with_Palittapongarnpim_full",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the full Palittapongarnpim et al. lineage for L1")
        parser.add_argument("-Shitikov",
                            "--with_Shitikov",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the Shitikov et al. lineage for L2")
        parser.add_argument("-Stucki",
                            "--with_Stucki",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the Stucki et al. lineage for L4")
        parser.add_argument("-IS",
                            "--with_IS",
                            type=self.__str2bool,
                            nargs='?',
                            const=True,
                            default=False,
                            help="Compute the number of various ISs")        
        parser.add_argument("--IS_evalue",
                            default='1e-7',
                            help="evalue when blasting Insertion Sequences",
                            type=str)
        parser.add_argument("--IS_prefix_size",
                            default=20,
                            help="size of the prefix in front of the IS",
                            type=int)

        # TODO: valeurs par défaut dans l'aide
        args = parser.parse_args()
        logging.basicConfig(level=args.loglevel)
        self._logger = logging.getLogger()
        if args.sra == "" and args.sra_list == "" and args.bioproject == "":
            sys.exit("No SRA found. Either 'sra', 'bioproject' or 'list' arguments must be provided.")
        if (args.sra != "" and args.sra_list != "")\
            or (args.sra != "" and args.bioproject != "")\
            or (args.bioproject != "" and args.sra_list != ""):
            sys.exit("Ambiguous arguments: only one of the arguments 'sra' and 'list_sra' can be passed. ")
        if args.sra != "":
            self.sras = [args.sra]
        elif args.list != "":
            with open(args.sra_list) as f:
                self.sras = f.read().split(os.linesep)
        else:
            self._results['bioproject'] = args.bioproject
        self.sras = [k for k in self.sras if k != '']
        self._output_dir = pathlib.Path(args.output_directory)
        self._output_dict = args.output_dict
        self._num_threads = args.num_threads
        self._spol_evalue = args.spol_evalue
        self._spol_nbmatches = args.spol_nbmatches
        self._with_spol_vitro = args.with_spol_vitro
        self._lineage_evalue = args.lineage_evalue
        self._lineage_window_size = args.lineage_blast_ws
        self._lineage_validate_size = args.lineage_validate_ws
        self._with_Coll = args.with_Coll
        self._with_Coll_full = args.with_Coll_full
        self._with_L6_animal = args.with_L6_animal
        self._with_PGG = args.with_PGG
        self._with_Palittapongarnpim = args.with_Palittapongarnpim
        self._with_Palittapongarnpim_full = args.with_Palittapongarnpim_full
        self._with_Shitikov = args.with_Shitikov
        self._with_Stucki = args.with_Stucki
        self._with_IS = args.with_IS
        self._IS_evalue = args.IS_evalue
        self._IS_prefix_size = args.IS_prefix_size



    def __bioproject_to_sra(self):
        self._logger.info(f"Collecting SRAs from bioproject{self._results['bioproject']}")
        ret = Bio.Entrez.esearch(db="sra",
                                 term=self._results['bioproject'],
                                 retmode="xml")
        dico = xmltodict.parse(ret.read())
        id_sras = dico['eSearchResult']['IdList']['Id']
        self.sras = []
        for k in id_sras:
            ret = Bio.Entrez.efetch(db="sra", id=k, retmode="xml")
            dico = xmltodict.parse(ret.read())
            self.sras.append(dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['@accession'])
            time.sleep(1)
    
    

    def _read_h37Rv(self):
        self._logger.info('Reading h37Rv reference genome')
        self._h37Rv = Bio.SeqIO.read(self._dir_data / 'NC_000962.3.fasta',
                                     'fasta')                         



    def _read_SIT(self):
        self._logger.info('Reading SIT database')
        wb = xlrd.open_workbook(self._dir_data / 'SIT.xls')
        ws = wb.sheet_by_index(0)
        self._spol_sit = {}
        for row in range(1, ws.nrows):
            spol, sit = ws.cell_value(row, 2).replace('n','\u25A0').replace('o','\u25A1'), ws.cell_value(row, 8)
            if spol not in self._spol_sit:
                self._spol_sit[spol] = sit



    def _prepare_directory(self):
        """
        Prepare the structure of sequence directory        
        """
        self._logger.debug('Preparing directory structure')
        self._dir = self._output_dir / self._sra
        self._dir.mkdir(exist_ok=True, parents=True)
        self._results[self._sra]['Directory'] = self._dir


    def _check_for_tools(self):

        
        
    def _check_for_tools_old(self):
        self._logger.info('Check for fastq-dump availability')
        local_fastqdump = list(filter(lambda x:x.startswith('fastq-dump'), os.listdir()))
        if local_fastqdump != []:
            self._logger.warning('fastq-dump found in local directory')
            self._fastq_dump = local_fastqdump[0]
            if sys.platform == 'win32':
                self._fastq_dump = 'fastq-dump.exe'
            elif sys.platform in ['darwin', 'linux']:
                self._fastq_dump = './fastq-dump'
        elif distutils.spawn.find_executable("fastq-dump.exe") != None:
            self._logger.warning('fastq-dump found in system directory')
            self._fastq_dump = 'fastq-dump.exe'
        elif distutils.spawn.find_executable("fastq-dump") != None:
            self._logger.warning('fastq-dump found in system directory')
            self._fastq_dump = 'fastq-dump'
        else:
            self._logger.warning('Trying to automatically download fastq-dump')
            if sys.platform == 'darwin':
                url = 'https://ftp-trace.ncbi.nlm.nih.gov/sra/sdk/2.10.7/sratoolkit.2.10.7-mac64.tar.gz'
            elif sys.platform == 'linux':
                url = 'https://ftp-trace.ncbi.nlm.nih.gov/sra/sdk/2.10.7/sratoolkit.2.10.7-ubuntu64.tar.gz'
            elif sys.platform == 'win32':
                url = 'https://ftp-trace.ncbi.nlm.nih.gov/sra/sdk/2.10.7/sratoolkit.2.10.7-win64.zip'
            else:
                sys.exit("fastq-dump (NCBI) is required")
            filename = url.rpartition('/')[2]
            urllib.request.urlretrieve(url, filename)
            if filename.endswith('gz'):
                tar = tarfile.open(filename, "r:gz")
                tar.extractall()
                tar.close()
            else:
                with zipfile.ZipFile(filename, 'r') as zip_ref:
                    zip_ref.extractall('.')
            filename = filename.replace('.zip', '').replace('.tar.gz','')
            p = pathlib.Path(filename) / 'bin'
            if sys.platform == 'win32':
                self._logger.warning('Downloading SRA files')
                shutil.copyfile(p / 'fastq-dump.exe', pathlib.Path('.'))
                self._fastq_dump = 'fastq-dump.exe'
            elif self._linux_or_mac:
                self._logger.warning('Downloading SRA files')
                shutil.copyfile(p / 'fastq-dump-orig.2.10.7', pathlib.Path('.') / 'fastq-dump')
                os.chmod('fastq-dump', stat.S_IEXEC)
                self._fastq_dump = './fastq-dump'
            else:
                sys.exit('Please install fastq-dump')
        self._logger.info('Check for makeblastdb availability')
        local_makeblastdb = list(filter(lambda x:x.startswith('makeblastdb'), os.listdir()))
        if local_makeblastdb != []:
            self._logger.warning('makeblastdb found in local directory')
            makeblastdb = local_makeblastdb[0]
            if sys.platform == 'win32':
                self._makebastdb = makeblastdb
                self._blastn = list(filter(lambda x:x.startswith('makeblastdb'), os.listdir()))[0]
            elif self._linux_or_mac:
                self._makebastdb = './'+makeblastdb        
                self._blastn = list(filter(lambda x:x.startswith('blastn'), os.listdir()))[0]
                self._blastn = './'+self._blastn
        elif distutils.spawn.find_executable("makeblastdb.exe") != None:
            self._logger.warning('makeblastdb found in system directory')
            self._makebastdb = 'makeblastdb.exe'
            self._blastn = 'blastn.exe'
        elif distutils.spawn.find_executable("makeblastdb") != None:
            self._logger.warning('makeblastdb found in system directory')
            self._makebastdb = 'makeblastdb'
            self._blastn = 'blastn'
        else:
            self._logger.warning('Trying to automatically download makeblastdb')
            if sys.platform == 'darwin':
                url = 'https://ftp.ncbi.nlm.nih.gov/blast/executables/blast+/2.10.0/ncbi-blast-2.10.0+-x64-macosx.tar.gz'
            elif sys.platform == 'linux':
                url = 'https://ftp.ncbi.nlm.nih.gov/blast/executables/blast+/2.10.0/ncbi-blast-2.10.0+-x64-linux.tar.gz'
            elif sys.platform == 'win32':
                url = 'https://ftp.ncbi.nlm.nih.gov/blast/executables/blast+/2.10.0/ncbi-blast-2.10.0+-x64-win64.tar.gz'
            else:
                sys.exit("makeblastdb (NCBI) is required")
            filename = url.rpartition('/')[2]
            urllib.request.urlretrieve(url, filename)
            tar = tarfile.open(filename, "r:gz")
            tar.extractall()
            tar.close()
            filename = filename.replace('.tar.gz','')
            p = pathlib.Path(filename) / 'ncbi-blast-2.10.0+' / 'bin'
            if sys.platform == 'win32':
                shutil.copyfile(p / 'makeblastdb.exe', pathlib.Path('.'))
                shutil.copyfile(p / 'blastn.exe', pathlib.Path('.'))
                self._blastn = 'blastn.exe'
                self._makebastdb = 'makeblastdb.exe'
            elif self._linux_or_mac:
                shutil.copyfile(p / 'makeblastdb', pathlib.Path('.') / 'makeblastdb')
                shutil.copyfile(p / self._blastn, pathlib.Path('.') / 'blastn')
                os.chmod('makeblastdb', stat.S_IEXEC)
                os.chmod('blastn', stat.S_IEXEC)
                self._blastn = './blastn'
                self._makebastdb = './makeblastdb'



    def _collect_sra(self):
        """
        Download the SRA files if needed.
        """
        self._sra_file_1 = self._dir / (self._sra + '_1')
        self._sra_file_1 = self._sra_file_1.with_suffix('.fasta')
        if not os.path.isfile(self._sra_file_1):
            sp.run([self._fastq_dump,
                    '--split-files',
                    '--fasta',
                    '-O', self._dir,
                    self._sra
                    ])
        self._sra_file_2 = self._dir / (self._sra + '_2')
        self._sra_file_2 = self._sra_file_2.with_suffix('.fasta')
        if not os.path.isfile(self._sra_file_1) or not os.path.isfile(self._sra_file_2):
            raise NotImplementedError("Fasta stem suffices are not _1 and _2")
        self._logger.warning('SRA files found in sequence directory')

         
        

    def _prepare_sra(self):
        """
        Merge the two downloaded fasta files of the paired Illumina WGS data.
        Valid only for linux or mac. For windows, we only take the first file.
        """
        if self._linux_or_mac:
            self._sra_shuffled = self._dir / (self._sra + '_shuffled')
            self._sra_shuffled = self._sra_shuffled.with_suffix('.fasta')
            if not os.path.isfile(self._sra_shuffled):
                self._logger.debug('Renaming reads')        
                for fic in ['_1', '_2']:
                    sp.run(["sed",
                            "-i", f"s/{self._sra}./{self._sra}{fic}./g",
                            (self._dir / (self._sra + fic)).with_suffix('.fasta')
                            ])
                self._logger.warning('Shuffling SRA files')      
                with open(self._sra_shuffled, "w+") as f:
                    sp.call(["cat",
                             self._sra_file_1,
                             self._sra_file_2],
                            stdout=f
                            )
        else:
            self._logger.info(f'We will only use one fasta file (OS={sys.platform})')
            self._sra_shuffled = self._sra_file_1


    def _seq_info(self):
        # We count the number of reads
        if self._linux_or_mac:
            proc1 = sp.Popen(["cat", self._sra_shuffled],
                             stdout=sp.PIPE)
            proc2 = sp.Popen(["grep", ">"],
                             stdin=proc1.stdout,
                             stdout=sp.PIPE)
            proc3 = sp.run(["wc", "-l"],
                             stdin=proc2.stdout,
                             stdout=sp.PIPE)
            self._results[self._sra]['nb_reads'] = int(proc3.stdout.decode('utf8'))
        else:
            with open(self._sra_shuffled) as f:
                self._results[self._sra]['nb_reads'] = f.read().count('>')
        self._logger.warning(f"{self._results[self._sra]['nb_reads']} reads found")
        # Getting read length
        head_seq = open(self._sra_shuffled).read(10000).split('>')[1]
        self._results[self._sra]['len_reads'] = len(''.join(head_seq.splitlines()[1:]))
        self._logger.warning(f"Read lengths: {self._results[self._sra]['len_reads']}")
        # Getting reads coverture
        exact_cov = self._results[self._sra]['nb_reads']*self._results[self._sra]['len_reads']/len(self._h37Rv)
        self._results[self._sra]['coverage'] = round(exact_cov, 2)
        self._logger.warning(f"Genome coverage: {self._results[self._sra]['coverage']}")
        if self._results[self._sra]['coverage'] < 50:
            self._logger.critical("Too low genomic coverage (< 50x)")
            


    def _ncbi_info(self):
        # TODO: biosample
        for key in ['center', 'date', 'location', 'name', 'strain', 'study', 'taxid']:
            self._results[self._sra][key] = ''
        if 'bioproject' not in self._results:
            self._results['bioproject'] = ''
        Bio.Entrez.email = ""
        ret = Bio.Entrez.efetch(db="sra", id=self._sra, retmode="xml")
        dico = xmltodict.parse(ret.read())
        # On ne gère (pour l'instant ?) que les ILLUMINA paired end
        if 'ILLUMINA' in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['PLATFORM']\
            and 'PAIRED' in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['DESIGN']['LIBRARY_DESCRIPTOR']['LIBRARY_LAYOUT']:
            # On récupère diverses informations
            try:
                attributes = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_ATTRIBUTES']['SAMPLE_ATTRIBUTE']
            except:
                return 
            location, date, center, strain = '', '', '', ''
            for k in attributes:
                if k['TAG'] == 'geographic location (country and/or sea)':
                    location = k['VALUE']
                elif k['TAG'] == 'collection date':
                    date = k['VALUE']
                elif k['TAG'] == 'INSDC center name':
                    center = k['VALUE']
                elif k['TAG'] == 'Strain':
                    strain = k['VALUE']
            dico0 = {'location': location,
                    'date': date,
                    'center': center,
                    'strain': strain,
                    'taxid' : dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME']['TAXON_ID'],
                    'name' : dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME']['SCIENTIFIC_NAME'],
                    'study' : dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['@alias']
                   }
            if isinstance(dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID'], list):
                dico0['bioproject'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID'][0]['#text']
            else:
                dico0['bioproject'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID']['#text']          
                
            dico0['center'] = ''
            try:
                dico0['center'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['@center_name']
            except:
                pass
        self._results[self._sra].update(dico0)


    def _make_blast_db(self):
        if not any([p.suffix == '.nin' for p in pathlib.Path(self._dir).iterdir()]):            
            completed = sp.run([self._makeblastdb,
                                '-in', self._sra_shuffled,
                                '-dbtype', 'nucl',
                                '-title', self._sra,
                                '-out', self._dir / self._sra])
            assert completed.returncode == 0


    def _blast_spoligo(self):
        self._logger.info('Blast of spacers')
        # TODO: refactor multiple spoligo blasts
        completed = sp.run([self._blastn,
                            '-num_threads', self._num_threads,
                            '-query', self._dir_data / 'spoligo_old.fasta',
                            '-evalue', self._spol_evalue,
                            '-task', "blastn",
                            '-db', self._dir / self._sra,
                            '-outfmt',
                            '10 qseqid sseqid sstart send qlen length score evalue',
                            '-out', self._dir / (self._sra + '_old-spol.blast')])
        assert completed.returncode == 0
        completed = sp.run([self._blastn,
                            '-num_threads', self._num_threads,
                            '-query', self._dir_data / 'spoligo_new.fasta',
                            '-evalue', self._spol_evalue,
                            '-task', "blastn",
                            '-db', self._dir / self._sra,
                            '-outfmt',
                            '10 qseqid sseqid sstart send qlen length score evalue',
                            '-out', self._dir / (self._sra + '_new-spol.blast')])
        assert completed.returncode == 0
        for index, item in enumerate(['_old', '_new']):
            self._results[self._sra]['spoligo'+['','_new'][index]] = ''
            with open(self._dir / (self._sra + item +'-spol.blast')) as f:
                matches = f.read()
                nb = open(self._dir_data / ('spoligo'+item+'.fasta')).read().count('>')
                for k in range(1,nb+1):
                    if matches.count('espaceur'+item[1:].capitalize()+str(k)+',') >= self._spol_nbmatches:
                        self._results[self._sra]['spoligo'+['','_new'][index]] += '\u25A0'
                    else:
                        self._results[self._sra]['spoligo'+['','_new'][index]] += '\u25A1'
            count_list = [matches.count('espaceur'+item[1:].capitalize()+str(k)+',') 
                          for k in range(1,nb+1)]
            self._results[self._sra]['spoligo'+['','_new'][index]+'_nb'] = count_list
        

    def _blast_spoligo_vitro(self):
        self._logger.info('Blast of in vitro spacers')
        completed = sp.run([self._blastn,
                            '-num_threads', self._num_threads,
                            '-query', self._dir_data / 'spoligo_vitro.fasta',
                            '-evalue', self._spol_evalue,
                            '-task', "blastn",
                            '-db', self._dir / self._sra,
                            '-outfmt',
                            '10 qseqid sseqid sstart send qlen length score evalue',
                            '-out', self._dir / (self._sra + '_vitro_old-spol.blast')])
        assert completed.returncode == 0
        completed = sp.run([self._blastn,
                            '-num_threads', self._num_threads,
                            '-query', self._dir_data / 'spoligo_vitro_new.fasta',
                            '-evalue', self._spol_evalue,
                            '-task', "blastn",
                            '-db', self._dir / self._sra,
                            '-outfmt',
                            '10 qseqid sseqid sstart send qlen length score evalue',
                            '-out', self._dir / (self._sra + '_vitro_new-spol.blast')])
        assert completed.returncode == 0
        if self._spol_vitro:
            # Spoligo old version
            with open(self._dir / (self._sra + '_vitro_old-spol.blast')) as f:
                matches = f.read()
                nb = int(open(self._dir_data / 'spoligo_vitro.fasta').read().count('>')/2)
                self._results[self._sra]['spoligo_vitro'] = ''
                for k in range(1,nb+1):
                    if min([matches.count('espaceur_vitroOld'+str(k)+','),matches.count('espaceur_vitroBOld'+str(k)+',')])/self._results[self._sra]['coverage'] > 0.05:
                        self._results[self._sra]['spoligo_vitro'] += '\u25A0'
                    else:
                        self._results[self._sra]['spoligo_vitro'] += '\u25A1'
            count_list = [(matches.count('espaceur_vitroOld'+str(k)+','),matches.count('espaceur_vitroBOld'+str(k)+',')) for k in range(1,nb+1)]
            self._results[self._sra]['spoligo_vitro_nb'] = count_list
            # Spoligo new version
            with open(self._dir / (self._sra + '_vitro_new-spol.blast')) as f:
                matches = f.read()
                nb = int(open(self._dir_data / 'spoligo_vitro_new.fasta').read().count('>')/2)
                self._results[self._sra]['spoligo_vitro_new'] = ''
                for k in range(1,nb+1):
                    if min([matches.count('espaceur_vitro_new'+str(k)+','),matches.count('espaceur_vitro_newB'+str(k)+',')])/self._results[self._sra]['coverage']>0.05:
                        self._results[self._sra]['spoligo_vitro_new'] += '\u25A0'
                    else:
                        self._results[self._sra]['spoligo_vitro_new'] += '\u25A1'
            count_list = [(matches.count('espaceur_vitro_new'+str(k)+','),matches.count('espaceur_vitro_newB'+str(k)+',')) for k in range(1,nb+1)]
            self._results[self._sra]['spoligo_vitro_new_nb'] = count_list
 

    def _SIT(self):
        self._logger.info('Getting SIT value')
        if self._results[self._sra]['spoligo'] in self._spol_sit:
            self._results[self._sra]['SIT'] = self._spol_sit[self._results[self._sra]['spoligo']]
        else:
            self._results[self._sra]['SIT'] = 'X'
        # SIT in vitro
        if self._with_spol_vitro:
            if self._results[self._sra]['spoligo_vitro'] in self._spol_sit:
                self._results[self._sra]['SIT_silico'] = self._spol_sit[self._results[self._sra]['spoligo_vitro']]
            else:
                self._results[self._sra]['SIT_silico'] = 'X'
   

    def __blast_lineages(self, Seq1, Seq2):
        if isinstance(Seq1, str):
            seq1, seq2 = Seq1, Seq2
        else:
            seq1, seq2 = str(Seq1.seq), str(Seq2.seq)
        with open(self._dir / 'snp.fasta','w') as f:
            f.write('>seq2\n'+seq2)
        result = sp.run([self._blastn,
                         "-num_threads", self._num_threads,
                         "-query", self._dir / 'snp.fasta',
                         "-evalue", self._lineage_evalue,
                         "-task", "blastn",
                         "-db", self._dir / self._sra,
                         "-outfmt", "10 sseq"],
                        stdout=sp.PIPE)
        formated_results = result.stdout.decode('utf8').splitlines()
        length = self._lineage_window_size
        lval = self._lineage_validate_size
        nb_seq1 = len([u for u in formated_results if seq1[length:length+lval] in u])
        nb_seq1 += len([u for u in formated_results if seq1[length-lval+1:length+1] in u])
        nb_seq2 = len([u for u in formated_results if seq2[length:length+lval] in u])
        nb_seq2 += len([u for u in formated_results if seq2[length-lval+1:length+1] in u])
        return nb_seq1, nb_seq2


    def _lineage_Coll(self):
        self._logger.info('Getting Coll lineage')
        wb = openpyxl.load_workbook(filename = self._dir_data / 'Coll.xlsx')
        ws = wb['Feuil1']
        lignee = []
        self._lineage_Coll_names = []
        length = self._lineage_window_size
        for row in ws.iter_rows(min_row=2):
            if row[1].value != None:
                pos = row[1].value-1
                assert self._h37Rv[pos] == row[3].value.split('/')[0]
                seq1 = self._h37Rv[pos-length:pos+length+1]
                if '*' not in row[0].value: 
                    seq2 = seq1[:length]+row[3].value.split('/')[1]+seq1[length+1:]
                else:
                    seq1 = seq1[:length]+row[3].value.split('/')[1]+seq1[length+1:]
                    seq2 = seq1[:length]+row[3].value.split('/')[0]+seq1[length+1:]
                nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
                lin = row[0].value.replace('lineage', '').replace('*','')
                self._lineage_Coll_names.append(lin)
                if nb_seq2>nb_seq1:
                    lignee.append(lin)
        lignee = sorted(set(lignee))
        self._results[self._sra]['lineage_Coll'] = lignee


    def _lineage_Coll_full(self):
        self._logger.info('Getting Coll complementary lineage')
        lignee = {}
        length = self._lineage_window_size
        with open(self._dir_data / 'Coll2.txt') as f:
            txt = f.read()
        for row in txt.split('\n')[1:-1]:
            lineage = row.split('\t')[0].lstrip('lineage')
            pos = int(row.split('\t')[1])-1
            ref_nt = row.split('\t')[4]
            assert self._h37Rv[pos] == ref_nt
            cible = row.split('\t')[5]
            seq1 = str(self._h37Rv[pos-length:pos+length+1].seq)
            seq2 = seq1[:length]+cible+seq1[length+1:]
            nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
            if nb_seq2>nb_seq1:
                if lineage not in lignee:
                    lignee[lineage] = [pos]
                else:
                    lignee[lineage].append(pos)
        #lignee = sorted(set(lignee))
        self._results[self._sra]['lineage_Coll_full'] = lignee


    def _lineage_L6_animal(self):
        self._logger.info('Getting L6+animal lineage')
        seq1 = 'ACGTCGATGGTCGCGACCTCCGCGGCATAGTCGAA'
        seq2 = "ACGTCGATGGTCGCGACTTCCGCGGCATAGTCGAA"
        with open(self._dir / 'snp.fasta','w') as f:
            f.write('>seq2\n'+seq2)
        result = sp.run([self._blastn,
                         "-num_threads", self._num_threads,
                         "-query", self._dir / 'snp.fasta',
                         "-evalue", self._lineage_evalue,
                         "-task", "blastn",
                         "-db", self._dir / self._sra,
                         "-outfmt", "10 sseq"],
                        stdout=sp.PIPE)
        formated_results = result.stdout.decode('utf8').splitlines()
        nb_seq1 = len([u for u in formated_results if seq1[13:18] in u])
        nb_seq1 += len([u for u in formated_results if seq1[17:22] in u])
        nb_seq2 = len([u for u in formated_results if seq2[13:18] in u])
        nb_seq2 += len([u for u in formated_results if seq2[17:22] in u])
        if nb_seq1 > nb_seq2:
            self._results[self._sra]['lineage_L6+animal'] = '1'
        elif nb_seq2 > nb_seq1:
            self._results[self._sra]['lineage_L6+animal'] = '2'
        else:
            self._results[self._sra]['lineage_L6+animal'] = 'X'


    def _lineage_PGG(self):
        self._logger.info('Getting PGG lineage')
        length = self._lineage_window_size
        lineage = []
        pos = 2154724
        seq1 = self._h37Rv[pos-length:pos+length+1]
        seq2 = seq1[:length-1]+'A'+seq1[length:]
        nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
        if nb_seq1 > nb_seq2:
            lineage.append('2')
        elif nb_seq2 > nb_seq1:
            lineage.append('1')
        else:
            lineage.append('X')
        pos = 7585-1
        seq1 = self._h37Rv[pos-length:pos+length+1]
        seq2 = seq1[:length]+'C'+seq1[length+1:]
        nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
        if nb_seq1 > nb_seq2:
            lineage.append('3')
        elif nb_seq2 > nb_seq1:
            lineage.append('1')
        else:
            lineage.append('X')
        self._results[self._sra]['lineage_PGG_cp'] = lineage
        if lineage == ['1', '1']:
            self._results[self._sra]['lineage_PGG'] = '1'
        elif lineage in [['1', '2'], ['2', '1']]:
            self._results[self._sra]['lineage_PGG'] = '2'
        elif lineage in [['2', '3'], ['3', '2']]:
            self._results[self._sra]['lineage_PGG'] = '3'
        else:
            self._results[self._sra]['lineage_PGG'] = 'X'


    def _lineage_Palittapongarnpim(self):
        """
        Refined L1 according to the first essential SNP per sublineage
        provided by Palittapongarnpim et al.
        """
        self._logger.info('Getting Palittapongarnpim lineage')
        length = self._lineage_window_size
        Lignee_Pali = {}
        wb = openpyxl.load_workbook(filename =  self._dir_data / 'Palittapongarnpim.xlsx')
        for feuille in wb.sheetnames:
            if 'lineage' in feuille:
                fiche = wb[feuille]
                for row in fiche.iter_rows(min_row=2):
                    if row[1].value != None and row[10].value == 'essential':
                        lignee = feuille.lstrip('lineage')
                        pos0 = row[1].value
                        source = row[3].value[0]
                        cible = row[3].value[2]
                        pos = pos0-1
                        assert self._h37Rv[pos] == source
                        seq1 = self._h37Rv[pos-length:pos+length+1]
                        seq2 = seq1[:length]+cible+seq1[length+1:]
                        Lignee_Pali[pos] = (str(seq1.seq), str(seq2.seq), lignee)
                        break
        lignee = []
        cpt = 1
        for item2,pos0 in enumerate(Lignee_Pali):
            seq1, seq2 = Lignee_Pali[pos0][:2]
            nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
            if nb_seq2>nb_seq1: 
                cpt += 1
                lignee.append(Lignee_Pali[pos0][2])
        lignee = [u for u in sorted(set(lignee))]
        self._results[self._sra]['lineage_Palittapongarnpim'] = lignee


    def _lineage_Palittapongarnpim_full(self):
        """
        Extended version of the refined L1, by considering all essential SNPs 
        per sublineage provided by Palittapongarnpim et al.
        """
        Lignee_Pali = {}
        par_lignee = {}
        length = self._lineage_window_size
        wb = openpyxl.load_workbook(filename =  self._dir_data / 'Palittapongarnpim.xlsx')
        for feuille in wb.sheetnames:
            if 'lineage' in feuille:
                fiche = wb[feuille]
                for row in fiche.iter_rows(min_row=2):
                    if row[1].value != None and row[10].value == 'essential':
                        lignee = feuille.lstrip('lineage')
                        pos0 = row[1].value
                        source = row[3].value[0]
                        cible = row[3].value[2]
                        pos = pos0-1
                        assert self._h37Rv[pos] == source
                        seq1 = self._h37Rv[pos-length:pos+length+1]
                        seq2 = seq1[:length]+cible+seq1[length+1:]
                        if lignee not in par_lignee:
                            par_lignee[lignee] = 1
                        else:
                            par_lignee[lignee] += 1
                        Lignee_Pali[pos] = (str(seq1.seq), str(seq2.seq), lignee)
        self._results[self._sra]['lineage_Pali_full'] = {}
        Lignee = []
        cpt = 1
        for item2,pos0 in enumerate(Lignee_Pali):
            seq1, seq2 = Lignee_Pali[pos0][:2]
            nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
            if nb_seq2>nb_seq1:
                cpt += 1
                Lignee.append([Lignee_Pali[pos0][2], pos0])
        lignee = [u for u in sorted([k[0] for k in Lignee])]
        for kk in sorted(set(lignee)):
            self._results[self._sra]['lineage_Pali_full'][kk] = [k[1] for k in Lignee if k[0]==kk]



    def _lineage_Shitikov(self):
        """
        L2 lineage refinement based on Shitikov list of SNPs.
        """
        length = self._lineage_window_size
        Lignee_Shitikov = {}
        wb4 = openpyxl.load_workbook(filename = self._dir_data / 'Shitikov.xlsx')
        fiche = wb4['Feuil1']
        for row in fiche.iter_rows(min_row=2):
            if row[1].value != None:
                lignee = row[0].value.lstrip('lineage').replace('*','')
                pos0 = row[1].value
                source = row[3].value[0]
                cible = row[3].value[2]
                pos = pos0-1
                assert self._h37Rv[pos] == source
                seq1 = self._h37Rv[pos-length:pos+length+1]
                seq2 = seq1[:20]+cible+seq1[21:]
                Lignee_Shitikov[pos] = (seq1, seq2, lignee)
        lignee = []
        cpt = 1
        for item2,pos0 in enumerate(Lignee_Shitikov):
            seq1, seq2 = Lignee_Shitikov[pos0][:2]
            nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
            if nb_seq2>nb_seq1:
                cpt += 1
                lignee.append(Lignee_Shitikov[pos0][2])
        lignee = [u for u in sorted(set(lignee))]
        self._results[self._sra]['lineage_Shitikov'] = lignee


    def _lineage_Stucki(self):
        """
        L4 lineage refinement based on Shitikov list of SNPs.
        """
        length = self._lineage_window_size
        Lignee_Stucki = {}
        wb5 = openpyxl.load_workbook(filename = self._dir_data / 'Stucki.xlsx')
        fiche = wb5['Feuil1']
        for row in fiche.iter_rows(min_row=2):
            if row[1].value != None:
                lignee = row[0].value.lstrip('lineage').replace('*','')
                pos0 = row[1].value
                source = row[3].value[0]
                cible = row[3].value[2]
                pos = pos0-1
                assert self._h37Rv[pos] == source
                seq1 = self._h37Rv[pos-length:pos+length+1]
                seq2 = seq1[:length]+cible+seq1[length+1:]
                Lignee_Stucki[pos] = (seq1, seq2, lignee)
        lignee = []
        cpt = 1
        for item2,pos0 in enumerate(Lignee_Stucki):
            seq1, seq2 = Lignee_Stucki[pos0][:2]
            nb_seq1, nb_seq2 = self.__blast_lineages(seq1, seq2)
            if nb_seq2>nb_seq1:
                cpt += 1
                lignee.append(Lignee_Stucki[pos0][2])
        lignee = [u for u in sorted(set(lignee))]
        if '4.10' in lignee:
            lignee.remove('4.10')
        else:
            lignee.append('4.10')
        self._results[self._sra]['lineage_Stucki'] = lignee


    def __similarity(self, x, y): 
        alignments = Bio.pairwise2.align.globalxx(x, y) 
        return alignments[0][2]/alignments[0][4] 


    def __get_position_in_h37Rv(self, name, seq):
        pos = 0
        try:
            pos = self._h37Rv.seq.index(seq)
        except ValueError:
            with open(self._dir / 'snp.fasta','w') as f:
                f.write(f'>{name}\n{seq}')
            # TODO: ranger data
            result = sp.run([self._blastn,
                             "-num_threads", self._num_threads,
                             "-query", self._dir / 'snp.fasta',
                             "-evalue", '1e-2',
                             "-task", "blastn",
                             "-db", self._dir_data / "h37Rv",
                             "-outfmt", "10 sstart"],
                            stdout=sp.PIPE)
            formated_results = result.stdout.decode('utf8').splitlines()
            if formated_results != []:
                pos = int(formated_results[0])
        return pos


    def _find_IS(self):
        p = pathlib.Path(self._dir_data) / 'IS'
        for IS in sorted(p.iterdir()):
            nom = IS.stem
            self._results[self._sra][nom] = []
            completed = sp.run([self._blastn,
                                '-num_threads', self._num_threads,
                                '-query', IS,
                                '-evalue', self._IS_evalue,
                                '-task', "blastn",
                                '-db', self._dir / self._sra,
                                "-max_target_seqs", "2000000", 
                                '-outfmt',
                                '10 qstart sstart send sseqid'],
                               stdout=sp.PIPE)
            assert completed.returncode == 0
            dd = completed.stdout.decode('utf8').splitlines()
            ee=[]
            for k in range(self._IS_prefix_size, self._results[self._sra]['len_reads']):
                ee.extend([u for u in dd if u.startswith('1,'+str(k)+',')])
            ee = [(u.split(',')[1], u.split(',')[3]) for u in ee if eval(u.split(',')[1])<eval(u.split(',')[2])]
            avant = []
            fasta_sequences = Bio.SeqIO.parse(self._sra_shuffled, 'fasta')
            if ee != []:
                for fasta in fasta_sequences:
                    name, sequence = fasta.id, str(fasta.seq)
                    for u in ee:
                        if name == u[1]:
                            seq = sequence[eval(u[0])-self._IS_prefix_size:eval(u[0])-1]
                            avant.append(seq)
            counter = collections.Counter(avant)
            if len(counter) > 0:
                self._logger.info(f'{nom} found in the following prefixes (with number of matches)')
                self._logger.info(", ".join([k+" ("+str(counter[k])+')' for k in counter]))
                if len(counter) == 1:
                    self._logger.warning(f"Infered number of {nom}: 1")
                    seq = next(iter(counter))
                    pos = self.__get_position_in_h37Rv(nom, seq)
                    self._results[self._sra][nom].append((seq, pos, counter[seq]))
                else:
                    L = []
                    for k in range(len(counter)-1): 
                        seq1 = list(counter.keys())[k] 
                        for l in range(k+1,len(counter)): 
                            seq2 = list(counter.keys())[l] 
                            L.append((seq1, seq2, self.__similarity(seq1,seq2))) 
                    todel = []
                    for k in counter:
                        try:
                            if counter[k] == 1 and max([u[2] for u in L if k in u])<0.9: 
                                todel.append(k)
                        except:
                            pass
                    for k in todel:
                        del counter[k]
                    L = [k for k in L if k[0] not in todel and k[1] not in todel]      
                    if len(L) != 0:    
                        G=nx.Graph()
                        G.add_weighted_edges_from(L)
                        G.remove_edges_from([(u,v) for u,v in G.edges if G[u][v]['weight'] <0.75])
                        comps = nx.connected_components(G)
                        cpt, s = 0, 'Final list of prefixes: '
                        for k in comps:
                            seq = max([l for l in k], key=counter.get)
                            nb = sum([counter[l] for l in k])
                            pos = self.__get_position_in_h37Rv(nom, seq)
                            s += f"{seq} ({pos}:{nb}), "  
                            self._results[self._sra][nom].append((seq, pos, nb))
                            cpt += 1
                        self._logger.warning(f"Inferred number of {nom}: {cpt}")
                        self._logger.info(s)
                    else:
                        self._logger.warning(f"Inferred number of {nom}: 0")
            else:
                self._logger.warning(f"Infrered number of {nom}: 0")
                    


    def _local_summary(self):
        """
        Final summary with generation of report files
        """
        self._logger.info('Writing results')
        print()
        print("======= SUMMARY =======")
        txt = ''
        for key in self._results[self._sra]:
            txt += f"{key}: {self._results[self._sra][key]}"
            txt += os.linesep
        txt += f'Executed in: {time.time()-self._t0:.2f} s.'
        print(txt)
        with open(self._dir / (self._sra + '.summary'), 'w') as f:
            f.write(txt)
        # Outputs a dictionary ?
        if self._output_dict:
            with open(self._dir / (self._sra + '.pkl'), 'wb') as f:
                pickle.dump(self._results[self._sra], f)
        try:
            os.remove(self._dir / 'snp.fasta')
        except:
            pass

    def _summary(self):        
        workbook = xlsxwriter.Workbook('results.xls')
        worksheet = workbook.add_worksheet()
        row, col = 0, 0
        col += 1
        merge_format = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'fg_color': 'yellow'})     
        if self._with_Coll:
            worksheet.merge_range(row, col, row, col+len(self._lineage_Coll_names)-1,
                                  'Coll et al. lineage',
                                  merge_format)
            for index, item in enumerate(self._lineage_Coll_names):
                worksheet.write_string(1, index+col, item, merge_format)
            col += len(self._lineage_Coll_names)

        row += 1
        col = 0
        worksheet.write(row, col, 'SRA')
        col += 1
        worksheet.write(row, col, 'Spoligotype in silico')        
        col += 1
        worksheet.write(row, col, 'SIT')        
        if self._with_Coll:
            col += len(self._lineage_Coll_names)
        worksheet.write(row, col, 'Bioproject')        

        row += 1
        col = 0
        for sra in self._results:
            worksheet.write_string(row, col, sra)
            col += 1
            worksheet.write_string(row, col, self._results[sra]['spoligo'])
            col += 1
            worksheet.write_string(row, col, self._results[sra]['SIT'])
            col += 1
            if self._with_Coll:
                for index, item in enumerate(self._lineage_Coll_names):
                    if item in self._results[sra]['lineage_Coll']:
                        worksheet.write_string(row, index+col, 'X')
                col += len(self._lineage_Coll_names)
            worksheet.write_string(row, col, self._results[sra]['bioproject'])
            row += 1    
            col = 0
        workbook.close()

   
    
if __name__ == "__main__":
    TBannotator()
